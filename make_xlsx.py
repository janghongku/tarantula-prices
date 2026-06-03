#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
가격현황 + 가격변동을 '한 엑셀 파일(여러 시트)'로 합친다.
  요약 / 전체현황 / 가격변동  3개 시트.

선행: make_report.py 가 종별_가격현황.csv, 가격변동_로그.csv 를 생성해 둠.
실행: python3 make_xlsx.py   →  거미_가격_아카이브.xlsx
(openpyxl 필요)
"""
import csv, os, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUTDIR = os.path.join(HERE, "outputs")
SNAP = os.path.join(OUTDIR, "종별_가격현황.csv")
LOG = os.path.join(OUTDIR, "가격변동_로그.csv")
NEWLOG = os.path.join(OUTDIR, "신규등록_로그.csv")
OUT = os.path.join(OUTDIR, "거미_가격_아카이브.xlsx")

MONEY = {"최초가", "현재가", "이전가", "변동가", "증감액", "가격"}
INTNUM = MONEY | {"증감률(%)", "변동횟수"}
WIDTH = {"거미": 26, "판매처": 13, "채널": 8, "최초관측일": 13, "최초가": 11, "현재가": 11,
         "변동횟수": 9, "가격추이": 42, "최근변동일": 13, "재고": 7, "날짜": 12, "이전가": 11, "변동가": 11,
         "증감액": 10, "증감률(%)": 10, "방향": 7, "등록일": 12, "가격": 11, "URL": 52}
HDR_FILL = PatternFill("solid", fgColor="3A2A16")
HDR_FONT = Font(bold=True, color="FFFFFF")
UP_FONT, DOWN_FONT = Font(color="C0392B", bold=True), Font(color="2471A3", bold=True)
THIN = Side(style="thin", color="E6E0D8")


def read_csv(path):
    if not os.path.exists(path):
        return [], []
    with open(path, encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    return (rows[0], rows[1:]) if rows else ([], [])


def add_table(wb, title, header, data, color_dir=False):
    ws = wb.create_sheet(title)
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(1, c)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center")
    dir_i = header.index("방향") + 1 if "방향" in header else None
    delta_i = header.index("증감액") + 1 if "증감액" in header else None
    for row in data:
        vals = []
        for h, v in zip(header, row):
            if h in INTNUM and v not in ("", None):
                try:
                    v = int(v)
                except ValueError:
                    pass
            vals.append(v)
        ws.append(vals)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(header) + 1):
            h = header[c - 1]; cell = ws.cell(r, c)
            if h in MONEY:
                cell.number_format = "#,##0"
        if color_dir and dir_i:
            d = ws.cell(r, dir_i).value
            f = UP_FONT if d == "인상" else DOWN_FONT if d == "인하" else None
            if f:
                ws.cell(r, dir_i).font = f
                if delta_i:
                    ws.cell(r, delta_i).font = f
    for c in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(c)].width = WIDTH.get(header[c - 1], 12)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
    return ws


def summary_sheet(wb, snap, log):
    sh, sd = snap; lh, ld = log
    ws = wb.create_sheet("요약")
    def idx(h, n): return h.index(n) if n in h else -1
    total = len(sd)
    instock = sum(1 for r in sd if idx(sh, "재고") >= 0 and r[idx(sh, "재고")] == "재고")
    species = len({r[idx(sh, "거미")] for r in sd}) if idx(sh, "거미") >= 0 else 0
    up = sum(1 for r in ld if idx(lh, "방향") >= 0 and r[idx(lh, "방향")] == "인상")
    down = sum(1 for r in ld if idx(lh, "방향") >= 0 and r[idx(lh, "방향")] == "인하")
    dates = [r[idx(lh, "날짜")] for r in ld if idx(lh, "날짜") >= 0 and r[idx(lh, "날짜")]]
    firstobs = [r[idx(sh, "최초관측일")] for r in sd if idx(sh, "최초관측일") >= 0 and r[idx(sh, "최초관측일")]]
    rows = [
        ("한국 타란튤라 판매가 아카이브", ""),
        ("생성 시각", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("전체 상품 수", total),
        ("재고 상품 수", instock),
        ("거미 종류 수(대략)", species),
        ("최초 관측 시작일", min(firstobs) if firstobs else "-"),
        ("", ""),
        ("누적 가격변동 건수", len(ld)),
        ("  ├ 인상", up),
        ("  └ 인하", down),
        ("가장 최근 변동일", max(dates) if dates else "-"),
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    for r in range(4, ws.max_row + 1):
        ws.cell(r, 1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24
    return ws


def main():
    snap = read_csv(SNAP)
    log = read_csv(LOG)
    newlog = read_csv(NEWLOG)
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    summary_sheet(wb, snap, log)
    if snap[0]:
        add_table(wb, "전체현황", snap[0], snap[1])
    if log[0]:
        add_table(wb, "가격변동", log[0], log[1], color_dir=True)
    if newlog[0]:
        add_table(wb, "신규입고", newlog[0], newlog[1])
    wb.save(OUT)
    print(f"엑셀 생성: {os.path.basename(OUT)}  (시트: 요약 / 전체현황 {len(snap[1])}행 / 가격변동 {len(log[1])}행 / 신규입고 {len(newlog[1])}행)")


if __name__ == "__main__":
    main()
